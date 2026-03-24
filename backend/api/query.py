# -*- coding: utf-8 -*-
"""
多源数据查询小程序版 - 查询API

Author: 飞书百万（AI助手）
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from typing import Optional, List, Any
from datetime import datetime
from pathlib import Path
import time
import json
import io
import os
import uuid
import logging
import re

from models import get_db_session, DatabaseConfig, QueryTemplate
from models.database import School, QueryLog
from services.user_service import UserService, QueryTemplateService, QueryLogService
from services.cache_service import get_query_cache
from core.security import get_current_user, TokenData
from core.config import settings
from core.sql_validator import validate_sql_for_miniapp, validate_sql_for_admin, SQLSecurityValidator
from db.connector import get_connector
from db.connection_manager import connection_manager


# ========== 错误信息脱敏 ==========
def sanitize_error(error: Exception) -> str:
    """
    脱敏错误信息，防止泄露数据库结构
    
    规则：
    - 隐藏表名、字段名等敏感信息
    - 隐藏具体路径、IP地址
    - 只保留通用错误描述
    """
    error_str = str(error)
    
    # 通用错误模式
    patterns = [
        # MySQL错误
        (r"Table '[^']+' doesn't exist", "表不存在"),
        (r"Unknown column '[^']+'", "字段不存在"),
        (r"Column '[^']+' in", "字段访问错误"),
        (r"Table '[^']\.\.\w+'", "表不存在"),
        
        # SQLServer错误
        (r"Invalid column name '[^']+'", "字段不存在"),
        (r"Invalid object name '[^']+'", "对象不存在"),
        
        # Oracle错误
        (r"ORA-\d+: .+", "数据库错误"),
        (r"ORA-\d+:", "数据库错误"),
        
        # 通用错误
        (r"connection refused", "连接失败"),
        (r"timeout", "连接超时"),
        (r"Access denied", "访问被拒绝"),
        (r"Unknown database", "数据库不存在"),
    ]
    
    for pattern, replacement in patterns:
        if re.search(pattern, error_str, re.IGNORECASE):
            return replacement
    
    # 如果匹配不到已知模式，返回通用错误
    # 但要移除可能泄露的敏感信息
    sanitized = error_str
    
    # 移除文件路径
    sanitized = re.sub(r'/[a-zA-Z]:/[^ ]+', '<path>', sanitized)
    sanitized = re.sub(r'/var/[^ ]+', '<path>', sanitized)
    sanitized = re.sub(r'C:\\[^ ]+', '<path>', sanitized)
    
    # 移除IP地址
    sanitized = re.sub(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', '<ip>', sanitized)
    
    # 移除端口号
    sanitized = re.sub(r':\d{2,5}(/|$)', ':<port>', sanitized)
    
    # 移除数据库名/用户名等敏感词
    sensitive_words = ['password', 'pwd', 'secret', 'token', 'key']
    for word in sensitive_words:
        sanitized = re.sub(rf'{word}=[^ ,]+', f'{word}=<hidden>', sanitized, flags=re.IGNORECASE)
    
    # 如果仍然包含敏感信息，返回通用错误
    if any(word in sanitized.lower() for word in ['table', 'column', 'field', 'mysql', 'oracle', 'sqlserver']):
        return "查询执行失败，请检查查询条件"
    
    # 截断过长错误
    if len(sanitized) > 100:
        sanitized = sanitized[:100] + "..."
    
    return sanitized


# ========== 数据脱敏配置 ==========
# 敏感字段模式（支持正则）
SENSITIVE_FIELD_PATTERNS = [
    r'password', r'pwd', r'passwd',
    r'secret', r'token', r'key', r'api_key',
    r'id_card', r'idcard', r'身份证',
    r'phone', r'mobile', r'电话', r'手机',
    r'email', r'邮箱',
    r'bank', r'银行', r'account', r'账号',
    r'address', r'地址',
]

# 脱敏规则
def mask_sensitive_value(value: Any, field_name: str) -> Any:
    """
    对敏感字段值进行脱敏处理
    
    Args:
        value: 原始值
        field_name: 字段名
        
    Returns:
        脱敏后的值
    """
    if value is None:
        return None
    
    # 检查是否为敏感字段
    field_lower = field_name.lower()
    is_sensitive = any(re.search(pattern, field_lower, re.IGNORECASE) 
                       for pattern in SENSITIVE_FIELD_PATTERNS)
    
    if not is_sensitive:
        return value
    
    value_str = str(value)
    
    # 根据字段类型进行不同脱敏
    if 'password' in field_lower or 'pwd' in field_lower or 'passwd' in field_lower:
        return '******'
    
    if 'phone' in field_lower or 'mobile' in field_lower or '手机' in field_lower:
        # 手机号：保留前3后4
        if len(value_str) == 11:
            return value_str[:3] + '****' + value_str[-4:]
        return '****'
    
    if 'id_card' in field_lower or 'idcard' in field_lower or '身份证' in field_lower:
        # 身份证：保留前4后4
        if len(value_str) >= 15:
            return value_str[:4] + '**********' + value_str[-4:]
        return '****'
    
    if 'email' in field_lower or '邮箱' in field_lower:
        # 邮箱：保留前缀首字符和域名
        if '@' in value_str:
            parts = value_str.split('@')
            if len(parts[0]) > 1:
                return parts[0][0] + '***@' + parts[1]
        return '***@***.***'
    
    if 'bank' in field_lower or '银行' in field_lower:
        # 银行卡：保留前4后4
        if len(value_str) >= 8:
            return value_str[:4] + '****' + value_str[-4:]
        return '****'
    
    if 'token' in field_lower or 'key' in field_lower or 'secret' in field_lower:
        # Token/Key：只显示前4个字符
        if len(value_str) > 4:
            return value_str[:4] + '****'
        return '****'
    
    # 默认脱敏：中间用*替代
    if len(value_str) > 4:
        return value_str[:2] + '****' + value_str[-2:]
    return '****'


def mask_sensitive_data(rows: List[dict], enabled: bool = True) -> List[dict]:
    """
    对查询结果进行数据脱敏
    
    Args:
        rows: 原始数据行
        enabled: 是否启用脱敏（默认启用）
        
    Returns:
        脱敏后的数据行
    """
    if not enabled or not rows:
        return rows
    
    masked_rows = []
    for row in rows:
        masked_row = {}
        for key, value in row.items():
            masked_row[key] = mask_sensitive_value(value, key)
        masked_rows.append(masked_row)
    
    return masked_rows


def log_error_safe(log_service, user_id: int, school_id: int, error: Exception, sql: str = ""):
    """
    安全记录错误日志（脱敏）
    """
    error_msg = sanitize_error(error)
    logger.warning(f"[Query Error] User {user_id}: {error_msg}")
    return error_msg

logger = logging.getLogger(__name__)

router = APIRouter(tags=["查询"])


# ========== 请求/响应模型 ==========

class SmartQueryRequest(BaseModel):
    """智能查询请求"""
    school_id: int = Field(..., description="学校ID")
    template_id: int = Field(..., description="查询模板ID")
    conditions: List[dict] = Field(default_factory=list, description="查询条件")
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    limit: Optional[int] = 100
    offset: Optional[int] = 0  # 分页偏移量


class BindingQueryRequest(BaseModel):
    """绑定查询请求（v1.2.0）"""
    binding_id: int = Field(..., description="功能绑定ID")
    conditions: List[dict] = Field(default_factory=list, description="查询条件")
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    limit: Optional[int] = 100
    offset: Optional[int] = 0  # 分页偏移量


class SqlQueryRequest(BaseModel):
    """SQL查询请求"""
    school_id: int
    sql: str
    database_id: Optional[int] = None  # 可选，指定数据库ID


class ExportRequest(BaseModel):
    """导出请求"""
    school_id: int
    template_id: Optional[int] = None
    binding_id: Optional[int] = None  # 新增：支持binding_id导出
    sql: Optional[str] = None
    conditions: List[dict] = Field(default_factory=list)
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    limit: Optional[int] = 10000  # 导出默认10000条
    export_all: bool = False  # 是否导出全部数据


# ========== 辅助函数 ==========

def check_user_school_permission(db: Session, user_id: int, school_id: int) -> bool:
    """检查用户是否有学校权限"""
    # 超级管理员有所有权限
    user = UserService.get_by_id(db, user_id)
    if user and user.role == 'admin':
        return True
    
    # 检查用户学校权限
    schools = UserService.get_user_schools(db, user_id)
    for school in schools:
        if school['id'] == school_id:
            return True
    return False


def generate_sql_from_template(template: QueryTemplate, conditions: List[dict],
                               start_time: str = None, end_time: str = None,
                               limit: int = None, offset: int = 0) -> str:
    """从模板生成SQL - 防SQL注入版本"""
    
    # ========== 处理 select_columns 返回字段 ==========
    sql = template.sql_template
    
    # 检查模板是否有 select_columns 配置
    select_columns = None
    if hasattr(template, 'select_columns') and template.select_columns:
        select_columns = template.select_columns
    elif hasattr(template, 'result_fields') and template.result_fields:
        select_columns = template.result_fields
    
    # 如果有 select_columns，替换 SQL 中的 SELECT * 
    if select_columns and isinstance(select_columns, list) and len(select_columns) > 0:
        # 格式: [{"column": "CUSTNAME", "alias": "姓名"}, ...]
        col_parts = []
        for col in select_columns:
            if isinstance(col, dict):
                col_name = col.get('column', '')
                col_alias = col.get('alias', '')
                if col_name:
                    if col_alias:
                        col_parts.append(f"{col_name} as '{col_alias}'")
                    else:
                        col_parts.append(col_name)
            elif isinstance(col, str):
                col_parts.append(col)
        
        if col_parts:
            # 替换 SELECT * 为实际字段
            import re
            sql = re.sub(r'SELECT\s+\*', 'SELECT ' + ', '.join(col_parts), sql, flags=re.IGNORECASE)
    
    # ========== 原有的安全检查逻辑 ==========
    
    # ========== 安全检查 ==========
    # 获取模板允许的字段白名单
    allowed_fields = set()
    if template.fields:
        for field in template.fields:
            # 兼容字典和对象两种格式
            if isinstance(field, dict):
                fn = field.get('field_name', '')
                if fn:  # 只添加非空字段名
                    allowed_fields.add(fn)
            else:
                try:
                    fn = field.field_name
                    if fn:
                        allowed_fields.add(fn)
                except:
                    pass
    
    # 如果白名单为空，则不启用字段过滤（兼容旧模板）
    enable_field_filter = bool(allowed_fields)
    
    # 允许的操作符白名单
    ALLOWED_OPERATORS = {'=', '!=', '<>', '>', '<', '>=', '<=', 'LIKE', 'NOT LIKE', 'IN', 'NOT IN', 'IS NULL', 'IS NOT NULL'}
    
    # 危险字符检测（用于值）
    def is_safe_value(value: str) -> bool:
        """检测值是否安全"""
        if not value:
            return True
        # 检测SQL注入特征
        dangerous_patterns = [
            ';', '--', '/*', '*/', ' xp_', ' sp_', '0x',
            'UNION', 'SELECT', 'INSERT', 'UPDATE', 'DELETE',
            'DROP', 'ALTER', 'EXEC', 'SCRIPT'
        ]
        value_upper = value.upper()
        for pattern in dangerous_patterns:
            if pattern in value_upper:
                return False
        return True
    
    # 处理条件
    where_clauses = []
    for cond in conditions:
        field = cond.get('field')
        operator = cond.get('operator', '=').upper()
        value = cond.get('value', '')
        
        if not field:
            continue
        
        # 验证字段名（白名单模式才验证）
        if enable_field_filter and field not in allowed_fields:
            logger.warning(f"[SQL Injection] 字段 {field} 不在白名单中")
            continue
        
        # 验证操作符
        if operator not in ALLOWED_OPERATORS:
            logger.warning(f"[SQL Injection] 操作符 {operator} 不允许")
            continue
        
        # 验证值安全性
        if not is_safe_value(value):
            logger.warning(f"[SQL Injection] 检测到危险值: {value[:50]}...")
            continue
        
        # 安全构建条件
        value_escaped = value.replace("'", "''")  # SQL转义
        
        if operator == 'LIKE':
            where_clauses.append(f"{field} LIKE '%{value_escaped}%'")
        elif operator == 'NOT LIKE':
            where_clauses.append(f"{field} NOT LIKE '%{value_escaped}%'")
        elif operator in ('IS NULL', 'IS NOT NULL'):
            where_clauses.append(f"{field} {operator}")
        elif operator in ('IN', 'NOT IN'):
            # 处理IN操作符，值用逗号分隔
            values = [v.strip().replace("'", "''") for v in value.split(',') if v.strip()]
            if values:
                values_str = "','".join(values)
                where_clauses.append(f"{field} {operator} ('{values_str}')")
        else:
            where_clauses.append(f"{field} {operator} '{value_escaped}'")
    
    # 添加时间条件（安全处理）
    if template.time_field:
        time_field = template.time_field
        if start_time:
            start_safe = start_time.replace("'", "''")
            where_clauses.append(f"{time_field} >= '{start_safe}'")
        if end_time:
            end_safe = end_time.replace("'", "''")
            where_clauses.append(f"{time_field} <= '{end_safe}'")
    
    # 组合WHERE
    if where_clauses:
        where_str = " AND ".join(where_clauses)
        if "WHERE" in sql.upper():
            sql = sql + " AND " + where_str
        else:
            sql = sql + " WHERE " + where_str
    
    # 添加LIMIT和OFFSET（防止全表扫描）
    if limit:
        limit = min(limit, 1000)  # 最大1000条
        sql = sql + f" LIMIT {limit}"
    elif template.default_limit:
        sql = sql + f" LIMIT {template.default_limit}"
    
    if offset > 0:
        sql = sql + f" OFFSET {offset}"
    
    return sql


# ========== 小程序端API ==========

@router.get("/user/schools")
async def get_user_schools(current_user: TokenData = Depends(get_current_user),
                          db: Session = Depends(get_db_session)):
    """
    获取用户授权的学校列表
    
    用户登录后获取自己有权访问的学校
    """
    # 超级管理员可以看到所有学校
    user = UserService.get_by_id(db, int(current_user.user_id))
    if user and user.role == 'admin':
        from models.database import School
        schools = db.query(School).filter(School.status == 'active').all()
        return {
            "code": 200,
            "message": "获取成功",
            "data": [s.to_dict() for s in schools]
        }
    
    # 普通用户获取授权学校
    schools = UserService.get_user_schools(db, int(current_user.user_id))
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": schools
    }


@router.get("/user/categories")
async def get_categories(school_id: int,
                        current_user: TokenData = Depends(get_current_user),
                        db: Session = Depends(get_db_session)):
    """
    获取学校的业务大类列表
    
    - **school_id**: 学校ID
    """
    # 检查权限
    if not check_user_school_permission(db, int(current_user.user_id), school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    categories = QueryTemplateService.get_categories(db, school_id)
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": categories
    }


@router.get("/user/templates")
async def get_templates(school_id: int, category: str = None,
                       current_user: TokenData = Depends(get_current_user),
                       db: Session = Depends(get_db_session)):
    """
    获取查询模板列表
    
    - **school_id**: 学校ID
    - **category**: 业务大类（可选）
    """
    # 检查权限
    if not check_user_school_permission(db, int(current_user.user_id), school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    if category:
        templates = QueryTemplateService.get_by_category(db, school_id, category)
    else:
        templates = QueryTemplateService.get_by_school(db, school_id)
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": [t.to_dict() for t in templates]
    }


@router.post("/user/query")
async def smart_query(request: SmartQueryRequest,
                     current_user: TokenData = Depends(get_current_user),
                     db: Session = Depends(get_db_session)):
    """
    智能查询
    
    - **school_id**: 学校ID
    - **template_id**: 查询模板ID
    - **conditions**: 查询条件
    - **start_time**: 开始时间
    - **end_time**: 结束时间
    - **limit**: 结果条数限制
    """
    user_id = int(current_user.user_id)
    
    # 检查权限
    if not check_user_school_permission(db, user_id, request.school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    # 获取模板
    template = db.query(QueryTemplate).filter(
        QueryTemplate.id == request.template_id,
        QueryTemplate.school_id == request.school_id
    ).first()
    
    if not template:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="查询模板不存在"
        )
    
    # 获取数据库配置
    # 从模板获取数据库配置ID或根据学校查找
    # 这里简化处理，假设模板关联了数据库配置
    from models.database import DatabaseConfig
    db_config = db.query(DatabaseConfig).filter(
        DatabaseConfig.school_id == request.school_id,
        DatabaseConfig.status == 'active'
    ).first()
    
    if not db_config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该学校未配置数据库"
        )
    
    # 生成SQL（带分页）
    sql = generate_sql_from_template(
        template,
        request.conditions,
        request.start_time,
        request.end_time,
        limit=request.limit,
        offset=request.offset
    )
    
    # 生成COUNT SQL（用于分页总数）
    # 移除 LIMIT 和 OFFSET 用于 COUNT
    sql_base = generate_sql_from_template(
        template,
        request.conditions,
        request.start_time,
        request.end_time,
        limit=None,
        offset=0
    )
    
    # 初始化查询开始时间
    start_time_ms = int(time.time() * 1000)
    
    # ========== SQL安全验证 ==========
    # 小程序端只允许执行SELECT查询，禁止任何修改/删除操作
    is_valid, error_msg = validate_sql_for_miniapp(sql, "smart_query")
    if not is_valid:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录非法SQL尝试
        logger.warning(
            f"[SQL Security] Smart query validation failed | "
            f"User: {user_id} | School: {request.school_id} | "
            f"Template: {template.name} | Error: {error_msg} | "
            f"SQL: {SQLSecurityValidator.sanitize_for_log(sql)}"
        )
        
        QueryLogService.create_log(
            db, user_id, request.school_id, template.id,
            template.name, {"conditions": request.conditions, "validation_failed": True},
            sql, 0, query_time, "failed", error_msg
        )
        
        return {
            "code": 403,
            "message": f"SQL安全验证失败: {error_msg}",
            "data": {
                "error": error_msg,
                "suggestion": "请联系管理员确认查询模板配置是否正确"
            }
        }
    # ========== SQL验证结束 ==========
    
    # ========== 查询缓存检查 ==========
    # 构建缓存键（不包含 offset/limit，因为分页数据可能需要重新查询）
    cache_config = {
        'db_type': db_config.db_type,
        'host': db_config.host if db_config.db_type != 'SQLite' else 'sqlite',
        'port': db_config.port if db_config.db_type != 'SQLite' else 0,
        'db_name': db_config.db_name,
    }
    cache = get_query_cache()
    cache_key_sql = sql_base  # 使用不带分页的SQL作为缓存键
    
    # 尝试从缓存获取总数
    cached_total = cache.get(cache_config, f"count:{cache_key_sql}")
    if cached_total is not None:
        logger.info(f"[Query] Cache hit for count query")
    
    # 尝试从缓存获取数据
    cached_result = cache.get(cache_config, sql)
    if cached_result is not None:
        logger.info(f"[Query] Cache hit for data query")
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 数据脱敏
        masked_rows = mask_sensitive_data(cached_result['rows'])
        
        return {
            "code": 200,
            "message": "查询成功(缓存)",
            "data": {
                "rows": masked_rows,
                "count": len(masked_rows),
                "total": cached_result.get('total', cached_total or 0),
                "offset": request.offset or 0,
                "limit": request.limit,
                "has_more": (cached_result.get('total', cached_total or 0)) > (request.offset or 0) + len(masked_rows),
                "query_time": query_time,
                "sql": sql,
                "from_cache": True
            }
        }
    # ========== 缓存检查结束 ==========
    
    # 执行查询
    try:
        # 构建数据库连接配置
        config = {
            'db_type': db_config.db_type,
            'db_name': db_config.db_name,
        }
        
        # 非SQLite数据库需要更多配置
        if db_config.db_type != 'SQLite':
            from core.security import decrypt_password
            
            # 尝试解密密码，如果失败则使用原始密码（兼容旧数据）
            try:
                plain_password = decrypt_password(db_config.password)
            except Exception:
                plain_password = db_config.password
            
            config.update({
                'host': db_config.host,
                'port': db_config.port,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            })
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        
        # 先查询数据
        result = executor.execute_query(sql)
        
        # 再查询总数（用于分页）
        total_count = cached_total or 0
        if cached_total is None:
            try:
                count_sql = f"SELECT COUNT(*) as total FROM ({sql_base}) as t"
                count_result = executor.execute_query(count_sql)
                if count_result and len(count_result) > 0:
                    total_count = count_result[0].get('total', 0) if hasattr(count_result[0], 'get') else count_result[0][0]
                    # 缓存总数
                    cache.set(cache_config, f"count:{cache_key_sql}", total_count, ttl=60)
            except Exception as count_err:
                logger.warning(f"COUNT查询失败: {count_err}")
                # COUNT失败不影响主查询
        
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 转换为字典
        rows = [dict(row) for row in result] if result else []
        
        # 缓存查询结果（不含脱敏数据）
        cache.set(cache_config, sql, {'rows': rows, 'total': total_count}, ttl=60)
        
        # 记录日志
        QueryLogService.create_log(
            db, user_id, request.school_id, template.id,
            template.name, {"conditions": request.conditions, "offset": request.offset, "limit": request.limit},
            sql, len(rows), query_time, "success"
        )
        
        # 数据脱敏
        masked_rows = mask_sensitive_data(rows)
        
        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "rows": masked_rows,
                "count": len(masked_rows),
                "total": total_count,
                "offset": request.offset or 0,
                "limit": request.limit,
                "has_more": total_count > (request.offset or 0) + len(masked_rows),
                "query_time": query_time,
                "sql": sql,
                "from_cache": False
            }
        }
        
    except Exception as e:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 构建详细错误信息（内部记录，保留详情）
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "sql": sql,
            "template_name": template.name,
            "suggestion": None
        }
        
        # 根据错误类型提供建议
        error_str = str(e).lower()
        if 'connection' in error_str or 'connect' in error_str:
            error_detail['suggestion'] = "数据库连接失败，请检查数据库配置和网络连接"
        elif 'syntax' in error_str or 'sql' in error_str:
            error_detail['suggestion'] = "SQL语法错误，请检查查询模板配置"
        elif 'permission' in error_str or 'denied' in error_str:
            error_detail['suggestion'] = "数据库权限不足，请联系管理员"
        elif 'timeout' in error_str:
            error_detail['suggestion'] = "查询超时，请缩小查询范围或添加时间条件"
        elif 'table' in error_str and ('not exist' in error_str or 'doesn\'t exist' in error_str):
            error_detail['suggestion'] = "表不存在，请检查模板配置或数据库结构"
        elif 'column' in error_str and ('unknown' in error_str or 'invalid' in error_str):
            error_detail['suggestion'] = "字段不存在，请检查模板配置或数据库结构"
        
        # 脱敏后的错误信息（返回给用户）
        safe_error_msg = sanitize_error(e)
        
        # 记录错误日志（内部保留详细错误信息）
        QueryLogService.create_log(
            db, user_id, request.school_id, template.id,
            template.name, {"conditions": request.conditions},
            sql, 0, query_time, "failed", str(e), error_detail
        )
        
        # 返回脱敏后的错误信息给用户
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"查询失败: {safe_error_msg}"
        )
        
        return {
            "code": 500,
            "message": f"查询失败: {str(e)}",
            "data": {
                "error": error_detail
            }
        }


@router.post("/user/query/binding", summary="绑定ID查询（v1.2.0）")
async def binding_query(request: BindingQueryRequest,
                        current_user: TokenData = Depends(get_current_user),
                        db: Session = Depends(get_db_session)):
    """
    通过binding_id进行查询（v1.2.0新API）
    
    - **binding_id**: 功能绑定ID
    - **conditions**: 查询条件
    - **start_time**: 开始时间
    - **end_time**: 结束时间
    - **limit**: 结果条数限制
    """
    from models.database import SchoolTemplateBinding, QueryTemplate, QueryField
    from sqlalchemy.orm import joinedload
    
    user_id = int(current_user.user_id)
    
    # 获取binding信息（使用id查询，同时预加载template和database关系）
    binding = db.query(SchoolTemplateBinding).options(
        joinedload(SchoolTemplateBinding.template),
        joinedload(SchoolTemplateBinding.database)
    ).filter(
        SchoolTemplateBinding.id == request.binding_id
    ).first()
    
    if not binding:
        raise HTTPException(
            status_code=404,
            detail="功能绑定不存在"
        )
    
    # 检查学校权限
    if not check_user_school_permission(db, user_id, binding.school_id):
        raise HTTPException(
            status_code=403,
            detail="无权限访问该学校"
        )
    
    # 获取模板
    template = binding.template
    if not template:
        raise HTTPException(
            status_code=404,
            detail="查询模板不存在"
        )
    
    # 验证模板必要字段
    if not template.sql_template:
        raise HTTPException(
            status_code=400,
            detail="模板SQL未配置"
        )
    if not template.table_name:
        raise HTTPException(
            status_code=400,
            detail="模板表名未配置"
        )
    
    # 获取数据库配置（优先使用已预加载的关系）
    db_config = binding.database if binding.database else db.query(DatabaseConfig).filter(
        DatabaseConfig.id == binding.database_id,
        DatabaseConfig.status == 'active'
    ).first()
    
    if not db_config:
        raise HTTPException(
            status_code=404,
            detail="数据库配置不存在或未启用"
        )
    
    try:
        # 记录查询开始时间
        start_time_ms = int(time.time() * 1000)
        
        # 初始化sql变量
        sql = ""
        
        # 使用通用SQL生成函数（带分页）
        sql = generate_sql_from_template(
            template,
            request.conditions,
            request.start_time,
            request.end_time,
            limit=request.limit,
            offset=request.offset
        )
        
        # 生成COUNT SQL（用于分页总数）
        sql_base = generate_sql_from_template(
            template,
            request.conditions,
            request.start_time,
            request.end_time,
            limit=None,
            offset=0
        )
        
        # 获取数据库连接（不区分大小写）
        db_type_lower = db_config.db_type.lower() if db_config.db_type else ''
        
        # 调试日志
        logger.info(f"数据库配置: db_type={db_config.db_type}, db_name={db_config.db_name}, host={db_config.host}")
        
        if db_type_lower == 'mysql':
            from core.security import decrypt_password
            plain_password = decrypt_password(db_config.password)
            config = {
                'db_type': db_config.db_type,  # 保持原始大小写给connector
                'host': db_config.host,
                'port': db_config.port or 3306,
                'db_name': db_config.db_name,
                'username': db_config.username,
                'password': plain_password
            }
        elif db_type_lower == 'oracle':
            from core.security import decrypt_password
            plain_password = decrypt_password(db_config.password)
            config = {
                'db_type': db_config.db_type,
                'host': db_config.host,
                'port': db_config.port or 1521,
                'db_name': db_config.db_name,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            }
        elif db_type_lower == 'sqlserver':
            from core.security import decrypt_password
            plain_password = decrypt_password(db_config.password)
            config = {
                'db_type': db_config.db_type,
                'host': db_config.host,
                'port': db_config.port or 1433,
                'db_name': db_config.db_name,
                'username': db_config.username,
                'password': plain_password
            }
        else:
            raise Exception(f"不支持的数据库类型: {db_config.db_type}")
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        
        # 先查询数据
        result = executor.execute_query(sql)
        
        # 再查询总数（用于分页）
        total_count = 0
        try:
            count_sql = f"SELECT COUNT(*) as total FROM ({sql_base}) as t"
            count_result = executor.execute_query(count_sql)
            if count_result and len(count_result) > 0:
                total_count = count_result[0].get('total', 0) if hasattr(count_result[0], 'get') else count_result[0][0]
        except Exception as count_err:
            logger.warning(f"COUNT查询失败: {count_err}")
            # COUNT失败不影响主查询
        
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录日志
        QueryLogService.create_log(
            db, user_id, binding.school_id, template.id,
            template.name, {"conditions": request.conditions, "binding_id": request.binding_id, "offset": request.offset, "limit": request.limit},
            sql, len(result), query_time, "success"
        )
        
        rows = [dict(row) for row in result] if result else []
        
        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "rows": rows,
                "count": len(rows),
                "total": total_count,
                "offset": request.offset or 0,
                "limit": request.limit,
                "has_more": total_count > (request.offset or 0) + len(rows),
                "query_time": query_time,
                "sql": sql
            }
        }
        
    except Exception as e:
        # 记录详细错误日志
        logger.error(f"查询失败: {str(e)}", exc_info=True)
        
        query_time = 0
        if 'start_time_ms' in locals():
            query_time = int(time.time() * 1000) - start_time_ms
        
        # 脱敏错误信息
        safe_error_msg = sanitize_error(e)
        
        QueryLogService.create_log(
            db, user_id, binding.school_id, template.id,
            template.name, {"conditions": request.conditions, "binding_id": request.binding_id},
            sql if 'sql' in locals() else "", 0, query_time, "failed", str(e)
        )
        
        raise HTTPException(
            status_code=500,
            detail=f"查询失败: {safe_error_msg}"
        )


@router.get("/user/history")
async def get_query_history(
    skip: int = 0, 
    limit: int = 30,
    school_id: int = None,
    status: str = None,
    start_date: str = None,
    end_date: str = None,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取查询历史（支持分页和筛选）
    
    - **skip**: 跳过条数
    - **limit**: 返回条数
    - **school_id**: 学校ID筛选（可选）
    - **status**: 状态筛选（success/failed）
    - **start_date**: 开始日期筛选（YYYY-MM-DD）
    - **end_date**: 结束日期筛选（YYYY-MM-DD）
    """
    user_id = int(current_user.user_id)
    
    # 管理员可以查看所有历史，普通用户只能看自己的
    query = db.query(QueryLog)
    if current_user.role != 'admin':
        query = query.filter(QueryLog.user_id == user_id)
    else:
        if user_id:
            query = query.filter(QueryLog.user_id == user_id)
    
    # 筛选条件
    if school_id:
        query = query.filter(QueryLog.school_id == school_id)
    if status:
        query = query.filter(QueryLog.status == status)
    if start_date:
        query = query.filter(QueryLog.created_at >= start_date)
    if end_date:
        query = query.filter(QueryLog.created_at <= end_date + " 23:59:59")
    
    # 分页
    total = query.count()
    history = query.order_by(QueryLog.created_at.desc()).offset(skip).limit(limit).all()
    
    # 获取学校名称
    result = []
    for h in history:
        item = h.to_dict()
        school = db.query(School).filter(School.id == h.school_id).first()
        item['school_name'] = school.name if school else None
        result.append(item)
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": result,
        "total": total
    }


@router.get("/user/history/{log_id}")
async def get_history_detail(
    log_id: int,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取历史记录详情
    
    - **log_id**: 历史记录ID
    """
    user_id = int(current_user.user_id)
    
    log = db.query(QueryLog).filter(QueryLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    # 权限检查
    if current_user.role != 'admin' and log.user_id != user_id:
        raise HTTPException(status_code=403, detail="无权限查看")
    
    # 获取学校名称
    result = log.to_dict()
    school = db.query(School).filter(School.id == log.school_id).first()
    result['school_name'] = school.name if school else None
    
    # 获取模板名称
    if log.template_id:
        template = db.query(QueryTemplate).filter(QueryTemplate.id == log.template_id).first()
        result['template_name'] = template.name if template else None
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": result
    }


@router.delete("/user/history/{log_id}")
async def delete_history(
    log_id: int,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    删除历史记录
    
    - **log_id**: 历史记录ID
    """
    user_id = int(current_user.user_id)
    
    log = db.query(QueryLog).filter(QueryLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    # 权限检查
    if current_user.role != 'admin' and log.user_id != user_id:
        raise HTTPException(status_code=403, detail="无权限删除")
    
    db.delete(log)
    db.commit()
    
    return {
        "code": 200,
        "message": "删除成功",
        "data": None
    }


@router.post("/user/sql")
async def direct_sql_query(
    request: SqlQueryRequest,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    直接执行SQL查询
    
    - **school_id**: 学校ID
    - **sql**: SQL语句（仅支持SELECT）
    """
    user_id = int(current_user.user_id)
    
    # 权限检查 - 普通用户需要学校权限
    if not check_user_school_permission(db, user_id, request.school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    # 检查是否是管理员（管理员可以执行SQL）
    user = UserService.get_by_id(db, user_id)
    if user.role != 'admin':
        # 非管理员只能执行查询，且需要配置允许
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="普通用户无直接SQL执行权限，请使用模板查询"
        )
    
    # ========== SQL安全验证 ==========
    # 管理员使用读写模式验证（允许SELECT/INSERT/UPDATE）
    is_valid, error_msg = validate_sql_for_admin(request.sql, "direct_sql_admin")
    if not is_valid:
        logger.warning(
            f"[SQL Security] Direct SQL validation failed | "
            f"User: {user_id} (admin) | School: {request.school_id} | "
            f"Error: {error_msg} | "
            f"SQL: {SQLSecurityValidator.sanitize_for_log(request.sql)}"
        )
        
        return {
            "code": 403,
            "message": f"SQL安全验证失败: {error_msg}",
            "data": {
                "error": error_msg
            }
        }
    # ========== SQL验证结束 ==========
    
    # 获取数据库配置
    if request.database_id:
        # 指定了数据库ID，使用指定的数据库
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.id == request.database_id,
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    else:
        # 未指定数据库ID，使用第一个可用的数据库
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    
    if not db_config:
        detail_msg = "指定的数据库不存在" if request.database_id else "该学校未配置数据库"
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=detail_msg
        )
    
    # 执行查询
    start_time_ms = int(time.time() * 1000)
    try:
        config = {
            'db_type': db_config.db_type,
            'db_name': db_config.db_name,
        }
        
        if db_config.db_type != 'SQLite':
            from core.security import decrypt_password
            try:
                plain_password = decrypt_password(db_config.password)
            except Exception:
                plain_password = db_config.password
            config.update({
                'host': db_config.host,
                'port': db_config.port,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            })
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        result = executor.execute_query(request.sql)
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录日志
        QueryLogService.create_log(
            db, user_id, request.school_id, None,
            "SQL查询", {"type": "direct_sql"},
            request.sql, len(result), query_time, "success"
        )
        
        # 转换为字典
        rows = [dict(row) for row in result] if result else []
        
        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "rows": rows,
                "count": len(rows),
                "query_time": query_time,
                "sql": request.sql
            }
        }
        
    except Exception as e:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 构建详细错误信息
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "sql": request.sql,
            "suggestion": None
        }
        
        # 根据错误类型提供建议
        error_str = str(e).lower()
        if 'connection' in error_str or 'connect' in error_str:
            error_detail['suggestion'] = "数据库连接失败，请检查数据库配置和网络连接"
        elif 'syntax' in error_str or 'sql' in error_str:
            error_detail['suggestion'] = "SQL语法错误，请检查SQL语句"
        elif 'permission' in error_str or 'denied' in error_str:
            error_detail['suggestion'] = "数据库权限不足，请联系管理员"
        elif 'timeout' in error_str:
            error_detail['suggestion'] = "查询超时，请优化SQL或缩小查询范围"
        elif 'table' in error_str and ('not exist' in error_str or 'doesn\'t exist' in error_str):
            error_detail['suggestion'] = "表不存在，请检查表名"
        elif 'column' in error_str and ('unknown' in error_str or 'invalid' in error_str):
            error_detail['suggestion'] = "字段不存在，请检查字段名"
        
        # 脱敏后的错误信息
        safe_error_msg = sanitize_error(e)
        
        # 记录错误日志（包含详细错误信息）
        QueryLogService.create_log(
            db, user_id, request.school_id, None,
            "SQL查询", {"type": "direct_sql"},
            request.sql, 0, query_time, "failed", str(e), error_detail
        )
        
        return {
            "code": 500,
            "message": f"查询失败: {safe_error_msg}",
            "data": {
                "error": error_detail
            }
        }


@router.post("/user/export")
async def export_query_result(
    request: ExportRequest,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    导出查询结果为Excel
    
    - **school_id**: 学校ID
    - **template_id**: 查询模板ID（可选）
    - **binding_id**: 功能绑定ID（v1.2.0推荐）
    - **sql**: SQL语句（可选，直接SQL查询时使用）
    - **conditions**: 查询条件
    - **start_time**: 开始时间
    - **end_time**: 结束时间
    - **limit**: 导出数量限制
    - **export_all**: 是否导出全部
    """
    user_id = int(current_user.user_id)
    
    # 权限检查
    if not check_user_school_permission(db, user_id, request.school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    # ========== 导出权限检查 ==========
    # 检查用户是否有导出权限
    from models.database import User, UserSchool
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 检查用户角色，管理员和普通用户都可导出
    # 如果需要限制导出功能和查询功能一致，可以复用 check_user_school_permission
    # 这里不做额外限制，但在日志中记录导出操作
    
    logger.info(f"[Export] User {user_id} exports data from school {request.school_id}")
    
    # 获取SQL语句
    sql = request.sql
    query_name = "数据导出"
    binding = None
    
    # 支持 binding_id 方式导出（v1.2.0）
    if request.binding_id:
        from models.database import SchoolTemplateBinding
        from sqlalchemy.orm import joinedload
        
        binding = db.query(SchoolTemplateBinding).options(
            joinedload(SchoolTemplateBinding.template),
            joinedload(SchoolTemplateBinding.database)
        ).filter(
            SchoolTemplateBinding.id == request.binding_id
        ).first()
        
        if not binding:
            raise HTTPException(status_code=404, detail="功能绑定不存在")
        
        template = binding.template
        if template:
            sql = generate_sql_from_template(
                template,
                request.conditions,
                request.start_time,
                request.end_time
            )
            query_name = template.name
            db_config = binding.database
    elif request.template_id:
        # 从模板生成SQL
        template = db.query(QueryTemplate).filter(
            QueryTemplate.id == request.template_id,
            QueryTemplate.school_id == request.school_id
        ).first()
        
        if not template:
            raise HTTPException(status_code=404, detail="查询模板不存在")
        
        sql = generate_sql_from_template(
            template,
            request.conditions,
            request.start_time,
            request.end_time
        )
        query_name = template.name
        
        # 获取数据库配置
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    
    if not sql:
        return {
            "code": 400,
            "message": "请提供SQL语句或模板ID/binding_id",
            "data": None
        }
    
    # 如果没有通过 binding_id 获取数据库配置，则需要单独获取
    if not db_config:
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    
    # ========== SQL安全验证 ==========
    # 导出功能只允许SELECT查询
    is_valid, error_msg = validate_sql_for_miniapp(sql, "export")
    if not is_valid:
        logger.warning(
            f"[SQL Security] Export SQL validation failed | "
            f"User: {user_id} | School: {request.school_id} | "
            f"Template: {request.template_id} | Error: {error_msg} | "
            f"SQL: {SQLSecurityValidator.sanitize_for_log(sql)}"
        )
        
        return {
            "code": 403,
            "message": f"SQL安全验证失败: {error_msg}",
            "data": {
                "error": error_msg,
                "suggestion": "导出功能只允许执行SELECT查询"
            }
        }
    # ========== SQL验证结束 ==========
    
    if not db_config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该学校未配置数据库"
        )
    
    # 执行查询
    start_time_ms = int(time.time() * 1000)
    try:
        config = {
            'db_type': db_config.db_type,
            'db_name': db_config.db_name,
        }
        
        if db_config.db_type != 'SQLite':
            from core.security import decrypt_password
            try:
                plain_password = decrypt_password(db_config.password)
            except Exception:
                plain_password = db_config.password
            config.update({
                'host': db_config.host,
                'port': db_config.port,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            })
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        
        # 限制导出数量（支持自定义和全部导出）
        limit_sql = sql
        if request.export_all:
            # 导出全部，不加限制
            pass
        elif request.limit and "LIMIT" not in sql.upper():
            limit_sql = sql + f" LIMIT {request.limit}"
        elif "LIMIT" not in sql.upper():
            limit_sql = sql + " LIMIT 10000"  # 默认10000
        
        result = executor.execute_query(limit_sql)
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        if not result:
            return {
                "code": 400,
                "message": "查询结果为空",
                "data": None
            }
        
        # 转换为字典列表
        rows = [dict(row) for row in result]
        
        # 生成Excel
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "查询结果"
            
            # 写入表头
            if rows:
                headers = list(rows[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # 写入数据
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        # 处理日期等特殊类型
                        if hasattr(value, 'isoformat'):
                            value = value.isoformat()
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 自动调整列宽
                for col_idx, header in enumerate(headers, 1):
                    max_length = len(str(header))
                    for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # 保存文件
            export_dir = Path("./exports")
            export_dir.mkdir(exist_ok=True)
            
            filename = f"export_{uuid.uuid4().hex[:8]}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = export_dir / filename
            
            wb.save(filepath)
            
            # 记录日志
            QueryLogService.create_log(
                db, user_id, request.school_id, request.template_id,
                query_name, {"type": "export", "filename": filename},
                sql, len(rows), query_time, "success"
            )
            
            return {
                "code": 200,
                "message": "导出成功",
                "data": {
                    "filename": filename,
                    "url": f"/exports/{filename}",
                    "count": len(rows),
                    "query_time": query_time
                }
            }
            
        except ImportError:
            return {
                "code": 500,
                "message": "Excel导出库未安装",
                "data": None
            }
        
    except Exception as e:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 构建详细错误信息
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "sql": sql,
            "suggestion": None
        }
        
        # 根据错误类型提供建议
        error_str = str(e).lower()
        if 'connection' in error_str or 'connect' in error_str:
            error_detail['suggestion'] = "数据库连接失败，请检查数据库配置和网络连接"
        elif 'syntax' in error_str or 'sql' in error_str:
            error_detail['suggestion'] = "SQL语法错误，请检查查询条件"
        elif 'permission' in error_str or 'denied' in error_str:
            error_detail['suggestion'] = "数据库权限不足，请联系管理员"
        elif 'timeout' in error_str:
            error_detail['suggestion'] = "查询超时，请缩小查询范围或添加时间条件"
        elif 'table' in error_str and ('not exist' in error_str or 'doesn\'t exist' in error_str):
            error_detail['suggestion'] = "表不存在，请检查模板配置"
        elif 'column' in error_str and ('unknown' in error_str or 'invalid' in error_str):
            error_detail['suggestion'] = "字段不存在，请检查模板配置"
        
        # 记录错误日志（包含详细错误信息）
        QueryLogService.create_log(
            db, user_id, request.school_id, request.template_id,
            query_name, {"type": "export"},
            sql, 0, query_time, "failed", str(e), error_detail
        )
        
        return {
            "code": 500,
            "message": f"导出失败: {str(e)}",
            "data": {
                "error": error_detail
            }
        }




# ========== 收藏相关 API ==========

class FavoriteRequest(BaseModel):
    """收藏请求"""
    binding_id: Optional[int] = None
    template_id: Optional[int] = None
    school_id: int
    query_name: str
    query_params: Optional[List[dict]] = Field(default_factory=list)
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    sort_fields: Optional[str] = None


@router.get("/user/favorites", summary="获取收藏列表")
async def get_favorites(
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """获取当前用户的收藏列表"""
    user_id = int(current_user.user_id)
    
    from models.database import QueryFavorite
    favorites = db.query(QueryFavorite).filter(
        QueryFavorite.user_id == user_id
    ).order_by(QueryFavorite.created_at.desc()).all()
    
    return {
        "code": 200,
        "data": [f.to_dict() for f in favorites]
    }


@router.post("/user/favorites", summary="添加收藏")
async def add_favorite(
    request: FavoriteRequest,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """添加查询收藏"""
    user_id = int(current_user.user_id)
    
    # 检查是否已存在相同收藏
    from models.database import QueryFavorite
    existing = db.query(QueryFavorite).filter(
        QueryFavorite.user_id == user_id,
        QueryFavorite.binding_id == request.binding_id,
        QueryFavorite.template_id == request.template_id,
        QueryFavorite.school_id == request.school_id
    ).first()
    
    if existing:
        # 更新现有收藏
        existing.query_name = request.query_name
        existing.query_params = request.query_params
        existing.start_time = request.start_time
        existing.end_time = request.end_time
        existing.sort_fields = request.sort_fields
        db.commit()
        
        return {
            "code": 200,
            "message": "收藏已更新",
            "data": existing.to_dict()
        }
    
    # 创建新收藏
    favorite = QueryFavorite(
        user_id=user_id,
        school_id=request.school_id,
        binding_id=request.binding_id,
        template_id=request.template_id,
        query_name=request.query_name,
        query_params=request.query_params,
        start_time=request.start_time,
        end_time=request.end_time,
        sort_fields=request.sort_fields
    )
    
    db.add(favorite)
    db.commit()
    db.refresh(favorite)
    
    return {
        "code": 200,
        "message": "收藏成功",
        "data": favorite.to_dict()
    }


@router.delete("/user/favorites/{favorite_id}", summary="删除收藏")
async def delete_favorite(
    favorite_id: int,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """删除收藏"""
    user_id = int(current_user.user_id)
    
    from models.database import QueryFavorite
    favorite = db.query(QueryFavorite).filter(
        QueryFavorite.id == favorite_id,
        QueryFavorite.user_id == user_id
    ).first()
    
    if not favorite:
        raise HTTPException(status_code=404, detail="收藏不存在")
    
    db.delete(favorite)
    db.commit()
    
    return {
        "code": 200,
        "message": "删除成功"
    }
