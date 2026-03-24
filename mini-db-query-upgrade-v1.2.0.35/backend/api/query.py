# -*- coding: utf-8 -*-
"""
多源数据查询小程序版 - 查询API

Author: 飞书百万（AI助手）
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from typing import Optional, List, Any
from datetime import datetime
from pathlib import Path
import time
import json
import io
import os
import uuid
import logging

from models import get_db_session, DatabaseConfig, QueryTemplate
from models.database import School, QueryLog
from services.user_service import UserService, QueryTemplateService, QueryLogService
from core.security import get_current_user, TokenData
from core.config import settings
from core.sql_validator import validate_sql_for_miniapp, validate_sql_for_admin, SQLSecurityValidator
from db.connector import get_connector
from db.connection_manager import connection_manager

logger = logging.getLogger(__name__)

router = APIRouter(tags=["查询"])


# ========== 请求/响应模型 ==========

class SmartQueryRequest(BaseModel):
    """智能查询请求"""
    school_id: int = Field(..., description="学校ID")
    template_id: int = Field(..., description="查询模板ID")
    conditions: List[dict] = Field(default_factory=list, description="查询条件")
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    limit: Optional[int] = 100
    offset: Optional[int] = 0  # 分页偏移量


class BindingQueryRequest(BaseModel):
    """绑定查询请求（v1.2.0）"""
    binding_id: int = Field(..., description="功能绑定ID")
    conditions: List[dict] = Field(default_factory=list, description="查询条件")
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    limit: Optional[int] = 100
    offset: Optional[int] = 0  # 分页偏移量


class SqlQueryRequest(BaseModel):
    """SQL查询请求"""
    school_id: int
    sql: str
    database_id: Optional[int] = None  # 可选，指定数据库ID


class ExportRequest(BaseModel):
    """导出请求"""
    school_id: int
    template_id: Optional[int] = None
    binding_id: Optional[int] = None  # 新增：支持binding_id导出
    sql: Optional[str] = None
    conditions: List[dict] = Field(default_factory=list)
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    limit: Optional[int] = 10000  # 导出默认10000条
    export_all: bool = False  # 是否导出全部数据


# ========== 辅助函数 ==========

def check_user_school_permission(db: Session, user_id: int, school_id: int) -> bool:
    """检查用户是否有学校权限"""
    # 超级管理员有所有权限
    user = UserService.get_by_id(db, user_id)
    if user and user.role == 'admin':
        return True
    
    # 检查用户学校权限
    schools = UserService.get_user_schools(db, user_id)
    for school in schools:
        if school['id'] == school_id:
            return True
    return False


def generate_sql_from_template(template: QueryTemplate, conditions: List[dict],
                               start_time: str = None, end_time: str = None,
                               limit: int = None, offset: int = 0) -> str:
    """从模板生成SQL"""
    sql = template.sql_template
    
    # 处理条件
    where_clauses = []
    for cond in conditions:
        field = cond.get('field')
        operator = cond.get('operator', '=')
        value = cond.get('value', '')
        
        if not value:
            continue
        
        if operator.upper() == 'LIKE':
            where_clauses.append(f"{field} LIKE '%{value}%'")
        else:
            where_clauses.append(f"{field} {operator} '{value}'")
    
    # 添加时间条件
    if template.time_field:
        if start_time:
            where_clauses.append(f"{template.time_field} >= '{start_time}'")
        if end_time:
            where_clauses.append(f"{template.time_field} <= '{end_time}'")
    
    # 组合WHERE
    if where_clauses:
        where_str = " AND ".join(where_clauses)
        if "WHERE" in sql.upper():
            sql = sql + " AND " + where_str
        else:
            sql = sql + " WHERE " + where_str
    
    # 添加LIMIT和OFFSET（支持分页）
    if limit:
        sql = sql + f" LIMIT {limit}"
    elif template.default_limit:
        sql = sql + f" LIMIT {template.default_limit}"
    
    if offset > 0:
        sql = sql + f" OFFSET {offset}"
    
    return sql


# ========== 小程序端API ==========

@router.get("/user/schools")
async def get_user_schools(current_user: TokenData = Depends(get_current_user),
                          db: Session = Depends(get_db_session)):
    """
    获取用户授权的学校列表
    
    用户登录后获取自己有权访问的学校
    """
    # 超级管理员可以看到所有学校
    user = UserService.get_by_id(db, int(current_user.user_id))
    if user and user.role == 'admin':
        from models.database import School
        schools = db.query(School).filter(School.status == 'active').all()
        return {
            "code": 200,
            "message": "获取成功",
            "data": [s.to_dict() for s in schools]
        }
    
    # 普通用户获取授权学校
    schools = UserService.get_user_schools(db, int(current_user.user_id))
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": schools
    }


@router.get("/user/categories")
async def get_categories(school_id: int,
                        current_user: TokenData = Depends(get_current_user),
                        db: Session = Depends(get_db_session)):
    """
    获取学校的业务大类列表
    
    - **school_id**: 学校ID
    """
    # 检查权限
    if not check_user_school_permission(db, int(current_user.user_id), school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    categories = QueryTemplateService.get_categories(db, school_id)
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": categories
    }


@router.get("/user/templates")
async def get_templates(school_id: int, category: str = None,
                       current_user: TokenData = Depends(get_current_user),
                       db: Session = Depends(get_db_session)):
    """
    获取查询模板列表
    
    - **school_id**: 学校ID
    - **category**: 业务大类（可选）
    """
    # 检查权限
    if not check_user_school_permission(db, int(current_user.user_id), school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    if category:
        templates = QueryTemplateService.get_by_category(db, school_id, category)
    else:
        templates = QueryTemplateService.get_by_school(db, school_id)
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": [t.to_dict() for t in templates]
    }


@router.post("/user/query")
async def smart_query(request: SmartQueryRequest,
                     current_user: TokenData = Depends(get_current_user),
                     db: Session = Depends(get_db_session)):
    """
    智能查询
    
    - **school_id**: 学校ID
    - **template_id**: 查询模板ID
    - **conditions**: 查询条件
    - **start_time**: 开始时间
    - **end_time**: 结束时间
    - **limit**: 结果条数限制
    """
    user_id = int(current_user.user_id)
    
    # 检查权限
    if not check_user_school_permission(db, user_id, request.school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    # 获取模板
    template = db.query(QueryTemplate).filter(
        QueryTemplate.id == request.template_id,
        QueryTemplate.school_id == request.school_id
    ).first()
    
    if not template:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="查询模板不存在"
        )
    
    # 获取数据库配置
    # 从模板获取数据库配置ID或根据学校查找
    # 这里简化处理，假设模板关联了数据库配置
    from models.database import DatabaseConfig
    db_config = db.query(DatabaseConfig).filter(
        DatabaseConfig.school_id == request.school_id,
        DatabaseConfig.status == 'active'
    ).first()
    
    if not db_config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该学校未配置数据库"
        )
    
    # 生成SQL（带分页）
    sql = generate_sql_from_template(
        template,
        request.conditions,
        request.start_time,
        request.end_time,
        limit=request.limit,
        offset=request.offset
    )
    
    # 生成COUNT SQL（用于分页总数）
    # 移除 LIMIT 和 OFFSET 用于 COUNT
    sql_base = generate_sql_from_template(
        template,
        request.conditions,
        request.start_time,
        request.end_time,
        limit=None,
        offset=0
    )
    
    # 初始化查询开始时间
    start_time_ms = int(time.time() * 1000)
    
    # ========== SQL安全验证 ==========
    # 小程序端只允许执行SELECT查询，禁止任何修改/删除操作
    is_valid, error_msg = validate_sql_for_miniapp(sql, "smart_query")
    if not is_valid:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录非法SQL尝试
        logger.warning(
            f"[SQL Security] Smart query validation failed | "
            f"User: {user_id} | School: {request.school_id} | "
            f"Template: {template.name} | Error: {error_msg} | "
            f"SQL: {SQLSecurityValidator.sanitize_for_log(sql)}"
        )
        
        QueryLogService.create_log(
            db, user_id, request.school_id, template.id,
            template.name, {"conditions": request.conditions, "validation_failed": True},
            sql, 0, query_time, "failed", error_msg
        )
        
        return {
            "code": 403,
            "message": f"SQL安全验证失败: {error_msg}",
            "data": {
                "error": error_msg,
                "suggestion": "请联系管理员确认查询模板配置是否正确"
            }
        }
    # ========== SQL验证结束 ==========
    
    # 执行查询
    try:
        # 构建数据库连接配置
        config = {
            'db_type': db_config.db_type,
            'db_name': db_config.db_name,
        }
        
        # 非SQLite数据库需要更多配置
        if db_config.db_type != 'SQLite':
            from core.security import decrypt_password
            
            # 尝试解密密码，如果失败则使用原始密码（兼容旧数据）
            try:
                plain_password = decrypt_password(db_config.password)
            except Exception:
                plain_password = db_config.password
            
            config.update({
                'host': db_config.host,
                'port': db_config.port,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            })
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        
        # 先查询数据
        result = executor.execute_query(sql)
        
        # 再查询总数（用于分页）
        total_count = 0
        try:
            count_sql = f"SELECT COUNT(*) as total FROM ({sql_base}) as t"
            count_result = executor.execute_query(count_sql)
            if count_result and len(count_result) > 0:
                total_count = count_result[0].get('total', 0) if hasattr(count_result[0], 'get') else count_result[0][0]
        except Exception as count_err:
            logger.warning(f"COUNT查询失败: {count_err}")
            # COUNT失败不影响主查询
        
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录日志
        QueryLogService.create_log(
            db, user_id, request.school_id, template.id,
            template.name, {"conditions": request.conditions, "offset": request.offset, "limit": request.limit},
            sql, len(result), query_time, "success"
        )
        
        # 转换为字典
        rows = [dict(row) for row in result] if result else []
        
        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "rows": rows,
                "count": len(rows),
                "total": total_count,
                "offset": request.offset or 0,
                "limit": request.limit,
                "has_more": total_count > (request.offset or 0) + len(rows),
                "query_time": query_time,
                "sql": sql
            }
        }
        
    except Exception as e:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 构建详细错误信息
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "sql": sql,
            "template_name": template.name,
            "suggestion": None
        }
        
        # 根据错误类型提供建议
        error_str = str(e).lower()
        if 'connection' in error_str or 'connect' in error_str:
            error_detail['suggestion'] = "数据库连接失败，请检查数据库配置和网络连接"
        elif 'syntax' in error_str or 'sql' in error_str:
            error_detail['suggestion'] = "SQL语法错误，请检查查询模板配置"
        elif 'permission' in error_str or 'denied' in error_str:
            error_detail['suggestion'] = "数据库权限不足，请联系管理员"
        elif 'timeout' in error_str:
            error_detail['suggestion'] = "查询超时，请缩小查询范围或添加时间条件"
        elif 'table' in error_str and ('not exist' in error_str or 'doesn\'t exist' in error_str):
            error_detail['suggestion'] = "表不存在，请检查模板配置或数据库结构"
        elif 'column' in error_str and ('unknown' in error_str or 'invalid' in error_str):
            error_detail['suggestion'] = "字段不存在，请检查模板配置或数据库结构"
        
        # 记录错误日志（包含详细错误信息）
        QueryLogService.create_log(
            db, user_id, request.school_id, template.id,
            template.name, {"conditions": request.conditions},
            sql, 0, query_time, "failed", str(e), error_detail
        )
        
        return {
            "code": 500,
            "message": f"查询失败: {str(e)}",
            "data": {
                "error": error_detail
            }
        }


@router.post("/user/query/binding", summary="绑定ID查询（v1.2.0）")
async def binding_query(request: BindingQueryRequest,
                        current_user: TokenData = Depends(get_current_user),
                        db: Session = Depends(get_db_session)):
    """
    通过binding_id进行查询（v1.2.0新API）
    
    - **binding_id**: 功能绑定ID
    - **conditions**: 查询条件
    - **start_time**: 开始时间
    - **end_time**: 结束时间
    - **limit**: 结果条数限制
    """
    from models.database import SchoolTemplateBinding, QueryTemplate, QueryField
    from sqlalchemy.orm import joinedload
    
    user_id = int(current_user.user_id)
    
    # 获取binding信息（使用id查询，同时预加载template和database关系）
    binding = db.query(SchoolTemplateBinding).options(
        joinedload(SchoolTemplateBinding.template),
        joinedload(SchoolTemplateBinding.database)
    ).filter(
        SchoolTemplateBinding.id == request.binding_id
    ).first()
    
    if not binding:
        raise HTTPException(
            status_code=404,
            detail="功能绑定不存在"
        )
    
    # 检查学校权限
    if not check_user_school_permission(db, user_id, binding.school_id):
        raise HTTPException(
            status_code=403,
            detail="无权限访问该学校"
        )
    
    # 获取模板
    template = binding.template
    if not template:
        raise HTTPException(
            status_code=404,
            detail="查询模板不存在"
        )
    
    # 验证模板必要字段
    if not template.sql_template:
        raise HTTPException(
            status_code=400,
            detail="模板SQL未配置"
        )
    if not template.table_name:
        raise HTTPException(
            status_code=400,
            detail="模板表名未配置"
        )
    
    # 获取数据库配置（优先使用已预加载的关系）
    db_config = binding.database if binding.database else db.query(DatabaseConfig).filter(
        DatabaseConfig.id == binding.database_id,
        DatabaseConfig.status == 'active'
    ).first()
    
    if not db_config:
        raise HTTPException(
            status_code=404,
            detail="数据库配置不存在或未启用"
        )
    
    try:
        # 记录查询开始时间
        start_time_ms = int(time.time() * 1000)
        
        # 初始化sql变量
        sql = ""
        
        # 使用通用SQL生成函数（带分页）
        sql = generate_sql_from_template(
            template,
            request.conditions,
            request.start_time,
            request.end_time,
            limit=request.limit,
            offset=request.offset
        )
        
        # 生成COUNT SQL（用于分页总数）
        sql_base = generate_sql_from_template(
            template,
            request.conditions,
            request.start_time,
            request.end_time,
            limit=None,
            offset=0
        )
        
        # 获取数据库连接（不区分大小写）
        db_type_lower = db_config.db_type.lower() if db_config.db_type else ''
        
        # 调试日志
        logger.info(f"数据库配置: db_type={db_config.db_type}, db_name={db_config.db_name}, host={db_config.host}")
        
        if db_type_lower == 'mysql':
            from core.security import decrypt_password
            plain_password = decrypt_password(db_config.password)
            config = {
                'db_type': db_config.db_type,  # 保持原始大小写给connector
                'host': db_config.host,
                'port': db_config.port or 3306,
                'db_name': db_config.db_name,
                'username': db_config.username,
                'password': plain_password
            }
        elif db_type_lower == 'oracle':
            from core.security import decrypt_password
            plain_password = decrypt_password(db_config.password)
            config = {
                'db_type': db_config.db_type,
                'host': db_config.host,
                'port': db_config.port or 1521,
                'db_name': db_config.db_name,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            }
        elif db_type_lower == 'sqlserver':
            from core.security import decrypt_password
            plain_password = decrypt_password(db_config.password)
            config = {
                'db_type': db_config.db_type,
                'host': db_config.host,
                'port': db_config.port or 1433,
                'db_name': db_config.db_name,
                'username': db_config.username,
                'password': plain_password
            }
        else:
            raise Exception(f"不支持的数据库类型: {db_config.db_type}")
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        
        # 先查询数据
        result = executor.execute_query(sql)
        
        # 再查询总数（用于分页）
        total_count = 0
        try:
            count_sql = f"SELECT COUNT(*) as total FROM ({sql_base}) as t"
            count_result = executor.execute_query(count_sql)
            if count_result and len(count_result) > 0:
                total_count = count_result[0].get('total', 0) if hasattr(count_result[0], 'get') else count_result[0][0]
        except Exception as count_err:
            logger.warning(f"COUNT查询失败: {count_err}")
            # COUNT失败不影响主查询
        
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录日志
        QueryLogService.create_log(
            db, user_id, binding.school_id, template.id,
            template.name, {"conditions": request.conditions, "binding_id": request.binding_id, "offset": request.offset, "limit": request.limit},
            sql, len(result), query_time, "success"
        )
        
        rows = [dict(row) for row in result] if result else []
        
        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "rows": rows,
                "count": len(rows),
                "total": total_count,
                "offset": request.offset or 0,
                "limit": request.limit,
                "has_more": total_count > (request.offset or 0) + len(rows),
                "query_time": query_time,
                "sql": sql
            }
        }
        
    except Exception as e:
        # 记录详细错误日志
        logger.error(f"查询失败: {str(e)}", exc_info=True)
        
        query_time = 0
        if 'start_time_ms' in locals():
            query_time = int(time.time() * 1000) - start_time_ms
        
        QueryLogService.create_log(
            db, user_id, binding.school_id, template.id,
            template.name, {"conditions": request.conditions, "binding_id": request.binding_id},
            sql if 'sql' in locals() else "", 0, query_time, "failed", str(e)
        )
        
        raise HTTPException(
            status_code=500,
            detail=f"查询失败: {str(e)}"
        )


@router.get("/user/history")
async def get_query_history(
    skip: int = 0, 
    limit: int = 30,
    school_id: int = None,
    status: str = None,
    start_date: str = None,
    end_date: str = None,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取查询历史（支持分页和筛选）
    
    - **skip**: 跳过条数
    - **limit**: 返回条数
    - **school_id**: 学校ID筛选（可选）
    - **status**: 状态筛选（success/failed）
    - **start_date**: 开始日期筛选（YYYY-MM-DD）
    - **end_date**: 结束日期筛选（YYYY-MM-DD）
    """
    user_id = int(current_user.user_id)
    
    # 管理员可以查看所有历史，普通用户只能看自己的
    query = db.query(QueryLog)
    if current_user.role != 'admin':
        query = query.filter(QueryLog.user_id == user_id)
    else:
        if user_id:
            query = query.filter(QueryLog.user_id == user_id)
    
    # 筛选条件
    if school_id:
        query = query.filter(QueryLog.school_id == school_id)
    if status:
        query = query.filter(QueryLog.status == status)
    if start_date:
        query = query.filter(QueryLog.created_at >= start_date)
    if end_date:
        query = query.filter(QueryLog.created_at <= end_date + " 23:59:59")
    
    # 分页
    total = query.count()
    history = query.order_by(QueryLog.created_at.desc()).offset(skip).limit(limit).all()
    
    # 获取学校名称
    result = []
    for h in history:
        item = h.to_dict()
        school = db.query(School).filter(School.id == h.school_id).first()
        item['school_name'] = school.name if school else None
        result.append(item)
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": result,
        "total": total
    }


@router.get("/user/history/{log_id}")
async def get_history_detail(
    log_id: int,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    获取历史记录详情
    
    - **log_id**: 历史记录ID
    """
    user_id = int(current_user.user_id)
    
    log = db.query(QueryLog).filter(QueryLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    # 权限检查
    if current_user.role != 'admin' and log.user_id != user_id:
        raise HTTPException(status_code=403, detail="无权限查看")
    
    # 获取学校名称
    result = log.to_dict()
    school = db.query(School).filter(School.id == log.school_id).first()
    result['school_name'] = school.name if school else None
    
    # 获取模板名称
    if log.template_id:
        template = db.query(QueryTemplate).filter(QueryTemplate.id == log.template_id).first()
        result['template_name'] = template.name if template else None
    
    return {
        "code": 200,
        "message": "获取成功",
        "data": result
    }


@router.delete("/user/history/{log_id}")
async def delete_history(
    log_id: int,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    删除历史记录
    
    - **log_id**: 历史记录ID
    """
    user_id = int(current_user.user_id)
    
    log = db.query(QueryLog).filter(QueryLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    # 权限检查
    if current_user.role != 'admin' and log.user_id != user_id:
        raise HTTPException(status_code=403, detail="无权限删除")
    
    db.delete(log)
    db.commit()
    
    return {
        "code": 200,
        "message": "删除成功",
        "data": None
    }


@router.post("/user/sql")
async def direct_sql_query(
    request: SqlQueryRequest,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    直接执行SQL查询
    
    - **school_id**: 学校ID
    - **sql**: SQL语句（仅支持SELECT）
    """
    user_id = int(current_user.user_id)
    
    # 权限检查 - 普通用户需要学校权限
    if not check_user_school_permission(db, user_id, request.school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    # 检查是否是管理员（管理员可以执行SQL）
    user = UserService.get_by_id(db, user_id)
    if user.role != 'admin':
        # 非管理员只能执行查询，且需要配置允许
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="普通用户无直接SQL执行权限，请使用模板查询"
        )
    
    # ========== SQL安全验证 ==========
    # 管理员使用读写模式验证（允许SELECT/INSERT/UPDATE）
    is_valid, error_msg = validate_sql_for_admin(request.sql, "direct_sql_admin")
    if not is_valid:
        logger.warning(
            f"[SQL Security] Direct SQL validation failed | "
            f"User: {user_id} (admin) | School: {request.school_id} | "
            f"Error: {error_msg} | "
            f"SQL: {SQLSecurityValidator.sanitize_for_log(request.sql)}"
        )
        
        return {
            "code": 403,
            "message": f"SQL安全验证失败: {error_msg}",
            "data": {
                "error": error_msg
            }
        }
    # ========== SQL验证结束 ==========
    
    # 获取数据库配置
    if request.database_id:
        # 指定了数据库ID，使用指定的数据库
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.id == request.database_id,
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    else:
        # 未指定数据库ID，使用第一个可用的数据库
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    
    if not db_config:
        detail_msg = "指定的数据库不存在" if request.database_id else "该学校未配置数据库"
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=detail_msg
        )
    
    # 执行查询
    start_time_ms = int(time.time() * 1000)
    try:
        config = {
            'db_type': db_config.db_type,
            'db_name': db_config.db_name,
        }
        
        if db_config.db_type != 'SQLite':
            from core.security import decrypt_password
            try:
                plain_password = decrypt_password(db_config.password)
            except Exception:
                plain_password = db_config.password
            config.update({
                'host': db_config.host,
                'port': db_config.port,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            })
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        result = executor.execute_query(request.sql)
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 记录日志
        QueryLogService.create_log(
            db, user_id, request.school_id, None,
            "SQL查询", {"type": "direct_sql"},
            request.sql, len(result), query_time, "success"
        )
        
        # 转换为字典
        rows = [dict(row) for row in result] if result else []
        
        return {
            "code": 200,
            "message": "查询成功",
            "data": {
                "rows": rows,
                "count": len(rows),
                "query_time": query_time,
                "sql": request.sql
            }
        }
        
    except Exception as e:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 构建详细错误信息
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "sql": request.sql,
            "suggestion": None
        }
        
        # 根据错误类型提供建议
        error_str = str(e).lower()
        if 'connection' in error_str or 'connect' in error_str:
            error_detail['suggestion'] = "数据库连接失败，请检查数据库配置和网络连接"
        elif 'syntax' in error_str or 'sql' in error_str:
            error_detail['suggestion'] = "SQL语法错误，请检查SQL语句"
        elif 'permission' in error_str or 'denied' in error_str:
            error_detail['suggestion'] = "数据库权限不足，请联系管理员"
        elif 'timeout' in error_str:
            error_detail['suggestion'] = "查询超时，请优化SQL或缩小查询范围"
        elif 'table' in error_str and ('not exist' in error_str or 'doesn\'t exist' in error_str):
            error_detail['suggestion'] = "表不存在，请检查表名"
        elif 'column' in error_str and ('unknown' in error_str or 'invalid' in error_str):
            error_detail['suggestion'] = "字段不存在，请检查字段名"
        
        # 记录错误日志（包含详细错误信息）
        QueryLogService.create_log(
            db, user_id, request.school_id, None,
            "SQL查询", {"type": "direct_sql"},
            request.sql, 0, query_time, "failed", str(e), error_detail
        )
        
        return {
            "code": 500,
            "message": f"查询失败: {str(e)}",
            "data": {
                "error": error_detail
            }
        }


@router.post("/user/export")
async def export_query_result(
    request: ExportRequest,
    current_user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db_session)
):
    """
    导出查询结果为Excel
    
    - **school_id**: 学校ID
    - **template_id**: 查询模板ID（可选）
    - **sql**: SQL语句（可选，直接SQL查询时使用）
    - **conditions**: 查询条件
    - **start_time**: 开始时间
    - **end_time**: 结束时间
    """
    user_id = int(current_user.user_id)
    
    # 权限检查
    if not check_user_school_permission(db, user_id, request.school_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限访问该学校"
        )
    
    # 获取SQL语句
    sql = request.sql
    query_name = "数据导出"
    binding = None
    
    # 支持 binding_id 方式导出（v1.2.0）
    if request.binding_id:
        from models.database import SchoolTemplateBinding, QueryTemplate, DatabaseConfig
        from sqlalchemy.orm import joinedload
        
        binding = db.query(SchoolTemplateBinding).options(
            joinedload(SchoolTemplateBinding.template),
            joinedload(SchoolTemplateBinding.database)
        ).filter(
            SchoolTemplateBinding.id == request.binding_id
        ).first()
        
        if not binding:
            raise HTTPException(status_code=404, detail="功能绑定不存在")
        
        template = binding.template
        if template:
            sql = generate_sql_from_template(
                template,
                request.conditions,
                request.start_time,
                request.end_time
            )
            query_name = template.name
            db_config = binding.database
    elif request.template_id:
        # 从模板生成SQL
        template = db.query(QueryTemplate).filter(
            QueryTemplate.id == request.template_id,
            QueryTemplate.school_id == request.school_id
        ).first()
        
        if not template:
            raise HTTPException(status_code=404, detail="查询模板不存在")
        
        sql = generate_sql_from_template(
            template,
            request.conditions,
            request.start_time,
            request.end_time
        )
        query_name = template.name
        
        # 获取数据库配置
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    
    if not sql:
        return {
            "code": 400,
            "message": "请提供SQL语句或模板ID/binding_id",
            "data": None
        }
    
    # 如果没有通过 binding_id 获取数据库配置，则需要单独获取
    if not db_config:
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.school_id == request.school_id,
            DatabaseConfig.status == 'active'
        ).first()
    
    # ========== SQL安全验证 ==========
    # 导出功能只允许SELECT查询
    is_valid, error_msg = validate_sql_for_miniapp(sql, "export")
    if not is_valid:
        logger.warning(
            f"[SQL Security] Export SQL validation failed | "
            f"User: {user_id} | School: {request.school_id} | "
            f"Template: {request.template_id} | Error: {error_msg} | "
            f"SQL: {SQLSecurityValidator.sanitize_for_log(sql)}"
        )
        
        return {
            "code": 403,
            "message": f"SQL安全验证失败: {error_msg}",
            "data": {
                "error": error_msg,
                "suggestion": "导出功能只允许执行SELECT查询"
            }
        }
    # ========== SQL验证结束 ==========
    
    if not db_config:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该学校未配置数据库"
        )
    
    # 执行查询
    start_time_ms = int(time.time() * 1000)
    try:
        config = {
            'db_type': db_config.db_type,
            'db_name': db_config.db_name,
        }
        
        if db_config.db_type != 'SQLite':
            from core.security import decrypt_password
            try:
                plain_password = decrypt_password(db_config.password)
            except Exception:
                plain_password = db_config.password
            config.update({
                'host': db_config.host,
                'port': db_config.port,
                'username': db_config.username,
                'password': plain_password,
                'service_name': db_config.service_name
            })
        
        connector = get_connector(config['db_type'], config)
        if not connector or not connector.connect():
            raise Exception("数据库连接失败")
        
        from db.query_executor import QueryExecutor
        executor = QueryExecutor(connector)
        
        # 限制导出数量（支持自定义和全部导出）
        limit_sql = sql
        if request.export_all:
            # 导出全部，不加限制
            pass
        elif request.limit and "LIMIT" not in sql.upper():
            limit_sql = sql + f" LIMIT {request.limit}"
        elif "LIMIT" not in sql.upper():
            limit_sql = sql + " LIMIT 10000"  # 默认10000
        
        result = executor.execute_query(limit_sql)
        connector.close()
        
        query_time = int(time.time() * 1000) - start_time_ms
        
        if not result:
            return {
                "code": 400,
                "message": "查询结果为空",
                "data": None
            }
        
        # 转换为字典列表
        rows = [dict(row) for row in result]
        
        # 生成Excel
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "查询结果"
            
            # 写入表头
            if rows:
                headers = list(rows[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # 写入数据
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        # 处理日期等特殊类型
                        if hasattr(value, 'isoformat'):
                            value = value.isoformat()
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 自动调整列宽
                for col_idx, header in enumerate(headers, 1):
                    max_length = len(str(header))
                    for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # 保存文件
            export_dir = Path("./exports")
            export_dir.mkdir(exist_ok=True)
            
            filename = f"export_{uuid.uuid4().hex[:8]}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = export_dir / filename
            
            wb.save(filepath)
            
            # 记录日志
            QueryLogService.create_log(
                db, user_id, request.school_id, request.template_id,
                query_name, {"type": "export", "filename": filename},
                sql, len(rows), query_time, "success"
            )
            
            return {
                "code": 200,
                "message": "导出成功",
                "data": {
                    "filename": filename,
                    "url": f"/exports/{filename}",
                    "count": len(rows),
                    "query_time": query_time
                }
            }
            
        except ImportError:
            return {
                "code": 500,
                "message": "Excel导出库未安装",
                "data": None
            }
        
    except Exception as e:
        query_time = int(time.time() * 1000) - start_time_ms
        
        # 构建详细错误信息
        error_detail = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "sql": sql,
            "suggestion": None
        }
        
        # 根据错误类型提供建议
        error_str = str(e).lower()
        if 'connection' in error_str or 'connect' in error_str:
            error_detail['suggestion'] = "数据库连接失败，请检查数据库配置和网络连接"
        elif 'syntax' in error_str or 'sql' in error_str:
            error_detail['suggestion'] = "SQL语法错误，请检查查询条件"
        elif 'permission' in error_str or 'denied' in error_str:
            error_detail['suggestion'] = "数据库权限不足，请联系管理员"
        elif 'timeout' in error_str:
            error_detail['suggestion'] = "查询超时，请缩小查询范围或添加时间条件"
        elif 'table' in error_str and ('not exist' in error_str or 'doesn\'t exist' in error_str):
            error_detail['suggestion'] = "表不存在，请检查模板配置"
        elif 'column' in error_str and ('unknown' in error_str or 'invalid' in error_str):
            error_detail['suggestion'] = "字段不存在，请检查模板配置"
        
        # 记录错误日志（包含详细错误信息）
        QueryLogService.create_log(
            db, user_id, request.school_id, request.template_id,
            query_name, {"type": "export"},
            sql, 0, query_time, "failed", str(e), error_detail
        )
        
        return {
            "code": 500,
            "message": f"导出失败: {str(e)}",
            "data": {
                "error": error_detail
            }
        }


